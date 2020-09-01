# 11 

# The Command Query Separation Principle states that our methods or function should either be:
# comands that perform an action and change the state of a system or
# queries that return an awnser to the caller and do not chage the state of the system.
# So our methods should either be comands or queries but not both.

import openpyxl

wb = openpyxl.load_workbook("transaction2.xlsx")

wb.create_sheet("Sheet3") # This is an example of a comand method because it is responsible to perform a task. The task of creating a sheet.
# As result of calling this method the state of our system changes, in this case the workbook, changes. Everey time we call this method it will create a new sheet.

sheet = wb["Sheet1"] # This is an example of a query method. We use it to access a given sheet.
print(sheet)

cell = sheet.cell(2, 3) # This is an example of a query method. We use it to access a given cell. However this method violates the Command Query Separation Principle.
print(cell)

for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
# When run this for loop it will return the value of cell in in column A and row 1 to 5. And for row 6 to 9 it will return "None"

sheet.append([1, 2, 3]) # And after the for loop with the ".cell()" method the last row will be row 10 and not row 5.
# Because the ".cell()" method created the empty cell with the value "None"

wb.save("transaction3.xlsx")

numbers = ["a", "b", "c"] # We can see the Command Query Separation Principle in action when trying to acess a index tha it is not in the list.
numbers[4] # When we try to access the index 4 in this list. It will raise an exception, because that index does not exits. Instead of creating a new item.
# Here we are querying our list, so we should not change its state.